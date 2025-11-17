from __future__ import annotations

import io
from typing import Any

from celery import shared_task
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import AsyncJob, Order, Product, Category, Supplier
from .storage import get_storage


def _parse_boolean_from_cell(value: Any) -> bool:
    """
    Normalize Excel cell value into a strict boolean.

    Accepted inputs:
    - Python bools
    - Numeric 1/0 (int, float, Decimal)
    - Strings: "true"/"false", "1"/"0" (case-insensitive, trimmed)
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "1"}:
            return True
        if normalized in {"false", "0"}:
            return False
    raise ValueError("Invalid status; must be true/false or 1/0")


@shared_task
def export_orders_task(job_id: str, filters: dict[str, Any] | None = None) -> None:
    job = AsyncJob.objects.get(pk=job_id)
    job.mark_running()
    try:
        qs = Order.objects.select_related("user").prefetch_related("items__product").all().order_by("-created_at")
        filters = filters or {}
        if v := filters.get("status"):
            qs = qs.filter(status=v)
        if v := filters.get("user_id"):
            qs = qs.filter(user_id=v)
        if date_from := filters.get("date_from"):
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to := filters.get("date_to"):
            qs = qs.filter(created_at__date__lte=date_to)

        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        headers = [
            "order_number",
            "status",
            "created_at",
            "user_id",
            "username",
            "item_product_id",
            "item_product_name",
            "item_quantity",
        ]
        ws.append(headers)
        for order in qs:
            for it in order.items.all():
                ws.append(
                    [
                        order.order_number,
                        order.status,
                        timezone.localtime(order.created_at).isoformat(),
                        order.user_id,
                        order.user.username,
                        it.product_id,
                        it.product.name_uz,
                        it.quantity,
                    ]
                )
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        storage = get_storage()
        url = storage.save_bytes(
            data,
            f"orders_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        job.mark_success(url)
    except Exception as e:  # pragma: no cover - simplify
        job.mark_failed(str(e))


@shared_task
def import_products_task(job_id: str, file_bytes: bytes) -> None:
    job = AsyncJob.objects.get(pk=job_id)
    job.mark_running()
    try:
        from django.db import transaction

        with transaction.atomic():
            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            # Strict header
            expected = ["name_uz", "name_ru", "category", "supplier", "image_url", "description", "status"]
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            header = [c.value for c in header_cells]
            if header != expected:
                raise ValueError(f"Invalid header. Expected {expected}, got {header}")

            actions: list[tuple[int, str, str, list[str]]] = []  # row, action, message, errors
            created, updated, skipped = 0, 0, 0

            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                errors: list[str] = []
                name_uz = (row[0].value or "").strip()
                name_ru = (row[1].value or "").strip()
                cat_name = (row[2].value or "").strip()
                sup_name = (row[3].value or "").strip()
                image_url = (row[4].value or "").strip()
                description = (row[5].value or "").strip()
                raw_status = row[6].value
                try:
                    status_val = _parse_boolean_from_cell(raw_status)
                except ValueError as exc:
                    errors.append(str(exc))

                if not name_uz:
                    errors.append("name_uz required")
                if not name_ru:
                    errors.append("name_ru required")
                if not cat_name:
                    errors.append("category required")
                if not sup_name:
                    errors.append("supplier required")

                if errors:
                    actions.append((idx, "skipped", "Validation errors", errors))
                    skipped += 1
                    continue

                category, _c = Category.objects.get_or_create(name_uz=cat_name, defaults={"name_ru": cat_name})
                supplier, _s = Supplier.objects.get_or_create(name=sup_name)

                obj, was_created = Product.objects.update_or_create(
                    name_uz=name_uz,
                    defaults={
                        "name_ru": name_ru,
                        "category": category,
                        "supplier": supplier,
                        "image_url": image_url,
                        "description": description,
                        "status": status_val,
                    },
                )
                if was_created:
                    created += 1
                    actions.append((idx, "created", f"Product {obj.id}", []))
                else:
                    updated += 1
                    actions.append((idx, "updated", f"Product {obj.id}", []))

        # Write detailed summary (outside transaction to avoid locking during file generation)
        summary_wb = Workbook()
        s = summary_wb.active
        s.title = "Summary"
        s.append(["row", "action", "message", "errors"])
        for row_num, action, message, errs in actions:
            s.append([row_num, action, message, "; ".join(errs)])
        totals = summary_wb.create_sheet("Totals")
        totals.append(["created", "updated", "skipped"])
        totals.append([created, updated, skipped])
        buf = io.BytesIO()
        summary_wb.save(buf)
        url = get_storage().save_bytes(buf.getvalue(), "import_products_summary.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        job.mark_success(url)
    except Exception as e:  # pragma: no cover - simplify
        job.mark_failed(str(e))
